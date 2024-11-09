from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
import openpyxl
import time
from datetime import datetime

# Load Excel data (row 6)
wb = openpyxl.load_workbook("C:\\Users\\yogan\\Downloads\\bidhanoo23-24.xlsx")
sheet = wb.active

aadhaar_number = sheet.cell(row=6, column=1).value  # Aadhaar number from column 1
sanction_date = sheet.cell(row=6, column=9).value  # KCC loan sanction date from column 9

# Set up Selenium WebDriver with the ChromeDriver
service = Service("C:\\Users\\yogan\\Downloads\\chromedriver.exe")  # Path to chromedriver.exe
driver = webdriver.Chrome(service=service)
driver.get("https://fasalrin.gov.in")  # Open the government portal

# Add login steps if necessary here

# Step 1: Go to loan application section
loan_application_link = driver.find_element(By.LINK_TEXT, 'Loan Application')
loan_application_link.click()

# Wait for financial year dropdown to appear
WebDriverWait(driver, 10).until(
    EC.presence_of_element_located((By.ID, 'Financial Year'))
)

# Step 2: Select financial year (2023-24)
financial_year_dropdown = Select(driver.find_element(By.ID, 'Financial Year'))
financial_year_dropdown.select_by_visible_text('2023-24')

# Input Aadhaar number from Excel
aadhaar_field = driver.find_element(By.ID, 'Aadhaar Number')
aadhaar_field.clear()
aadhaar_field.send_keys(aadhaar_number)
aadhaar_field.send_keys(Keys.RETURN)

# Step 3: Choose account number from dropdown after Aadhaar
WebDriverWait(driver, 10).until(
    EC.presence_of_element_located((By.ID, 'Account numbers'))
)
account_number_dropdown = Select(driver.find_element(By.ID, 'Account numbers'))
account_number_dropdown.select_by_index(1)  # Choose the first account in the dropdown
driver.find_element(By.ID, 'OK').click()

# Step 4: Select application type as "Normal"
WebDriverWait(driver, 10).until(
    EC.presence_of_element_located((By.ID, 'Application Type'))
)
application_type_dropdown = Select(driver.find_element(By.ID, 'Application Type'))
application_type_dropdown.select_by_visible_text('Normal')


# Step 5: Check if DOB is present, if not calculate it based on age
try:
    dob_field = driver.find_element(By.ID, 'Date Of Birth (DD/MM/YYYY)')
    if dob_field.get_attribute('value') == "":
        # No DOB, calculate from age
        age = driver.find_element(By.ID, 'Age').text
        current_year = datetime.now().year
        dob_year = current_year - int(age)
        dob = f"01/04/{dob_year}"  # Set DOB as January 1st of calculated year
        dob_field.clear()
        dob_field.send_keys(dob)
    driver.find_element(By.ID, 'UPDATE & CONTINUE').click()
except:
    pass

# Step 6: Click Save and Continue on the next page
WebDriverWait(driver, 10).until(
    EC.element_to_be_clickable((By.ID, 'UPDATE & CONTINUE'))
).click()

# Step 7: Handle KCC loan sanctioned date based on Excel data
WebDriverWait(driver, 10).until(
    EC.presence_of_element_located((By.ID, 'KCC loan sanctioned / KCC renewed on'))
)

if sanction_date.year < 2018:
    kcc_date = "01/04/2018"  # Set KCC date to April 1st, 2018
else:
    kcc_date = sanction_date.strftime('%d/%m/%Y')  # Use the exact date from Excel

kcc_loan_date_field = driver.find_element(By.ID, 'KCC loan sanctioned / KCC renewed on')
kcc_loan_date_field.clear()
kcc_loan_date_field.send_keys(kcc_date)

# Stop the process after this
print("Process completed for row 6.")

# Close the browser
driver.quit()
